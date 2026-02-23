from sheets.base2 import BaseSheet2
from core.interfaces import SheetProcessor
from openpyxl.styles import Alignment, Side, Border 
from utils.dates import format_month_label_german
import re
import copy

# The excel columns A or B getting converted into numbers
def excel_col_to_number(col: str) -> int:
    num = 0
    for c in col:
        num = num * 26 + (ord(c) - ord("A") + 1)
    return num

# Changes the Numbers back into excel columns 
def number_to_excel_col(n: int) -> str:
    result = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        result = chr(r + ord("A")) + result
    return result

# Increases the Excel column by one
def next_column(col: str) -> str:
    return number_to_excel_col(excel_col_to_number(col) + 1)

# Increases the Excel column by two
def next_two_columns(col: str) -> str:
    return number_to_excel_col(excel_col_to_number(col) + 2)

# Reads an existing formula and increases the columns by one: =Service_Availability!AL27 -> =Service_Availability!AM27
def change_formula_one_column(formula: str) -> str:
    def repl(match):
        return next_column(match.group(1))
    
    # Pattern so the formula operator doesn't get interpreted as column
    pattern = r"([A-Z]+)(?=\d)"
    
    return re.sub(pattern, repl, formula)

# Reads an existing formula and increases the columns by two: =Volume_Report!CB12 -> =Volume_Report!CD12
def change_formula_two_columns(formula: str) -> str:
    def repl(match):
        return next_two_columns(match.group(1))
    
    # Pattern so the formula operator doesn't get interpreted as column
    pattern = r"([A-Z]+)(?=\d)"
    
    return re.sub(pattern, repl, formula)


class Dashboard(SheetProcessor):

    def process(self):
        # --- read config ---
        cfg = self.config["sheets"]["Dashboard"]

        # --- get worksheet ---
        ws = self.workbook[cfg["sheet_name"]]

        # --- base helper ---
        base = BaseSheet2(ws, cfg, self.input)

        # Step 1: Update report period dates in C16 and D16
        # Format: dd.mm.YYYY
        date_start_formatted = self.input["report_period"]["date_start"].strftime("%d.%m.%Y")
        date_end_formatted = self.input["report_period"]["date_end"].strftime("%d.%m.%Y")
        
        ws.cell(row=cfg["report_period"]["start_row"], column=cfg["report_period"]["start_col"]).value = date_start_formatted
        ws.cell(row=cfg["report_period"]["end_row"], column=cfg["report_period"]["end_col"]).value = date_end_formatted

        # Step 2: Process Table 1 (rows 20-23)
        table1_cfg = cfg["table1"]
        
        # First: Shift month rows (skip year row)
        new_col = base.delete_and_shift_columns(
            start_row=table1_cfg["start_row"],
            end_row=table1_cfg["end_row"],
            year_row=table1_cfg["year_row"],
            start_col=cfg["start_column"],
            end_col=cfg["end_column"]
        )
        
        # Update month label in month row for Table 1
        ws.cell(
            row=table1_cfg["month_row"],
            column=new_col
        ).value = format_month_label_german(self.input["report_period"]["date_start"])
        
        # Update formulas in Table 1 (shift by 1 column)
        for row in table1_cfg["formula_rows"]:
            src = ws.cell(row=row, column=new_col).value
            if isinstance(src, str) and src.startswith("="):
                ws.cell(row=row, column=new_col).value = change_formula_one_column(src)

        # Then: Handle year row (unmerge, shift, write, merge)
        self._handle_year_row(ws, table1_cfg, cfg["start_column"], cfg["end_column"])

        # Step 3: Process Table 2 (rows 26-29)
        table2_cfg = cfg["table2"]
        
        # First: Shift month rows (skip year row)
        new_col = base.delete_and_shift_columns(
            start_row=table2_cfg["start_row"],
            end_row=table2_cfg["end_row"],
            year_row=table2_cfg["year_row"],
            start_col=cfg["start_column"],
            end_col=cfg["end_column"]
        )
        
        # Update month label in month row for Table 2
        ws.cell(
            row=table2_cfg["month_row"],
            column=new_col
        ).value = format_month_label_german(self.input["report_period"]["date_start"])
        
        # Update formulas in Table 2 (shift by 1 column)
        for row in table2_cfg["formula_rows"]:
            src = ws.cell(row=row, column=new_col).value
            if isinstance(src, str) and src.startswith("="):
                ws.cell(row=row, column=new_col).value = change_formula_one_column(src)

        # Then: Handle year row (unmerge, shift, write, merge)
        self._handle_year_row(ws, table2_cfg, cfg["start_column"], cfg["end_column"])

    def _handle_year_row(self, ws, table_cfg, start_col, end_col):
        """
        Handles the year row by dynamically managing year labels across months.
    
        This function:
        1. Unmerges all existing year merges in the year row
        2. Determines the position of the current year based on the month being added
        3. Shifts the new year forward (if present) and writes year values
        4. Re-merges year cells correctly over their corresponding months
        5. Preserves cell styles during the process
    
        Logic:
        - Column C (start_col) always contains the old year (e.g., 2025 in a 2025->2026 transition)
        - Column N (end_col) is the rightmost column for months
        - The new year position shifts forward each month: (end_col - input_month + 2)
          Example: For Feb (month=2), new year is at column M (14-2+2=14)
                   For Mar (month=3), new year is at column L (14-3+2=13)
        - When input_month == 12 (December), the old year in column C is overwritten with the new year
        - Each month, if a new year exists, it shifts one column forward and merges to column N
    
        Example transition (2025 -> 2026):
        - Month 1 (Jan): 2025 merged C-M, 2026 written in N (single cell)
        - Month 2 (Feb): 2025 merged C-L, 2026 at M merged M-N
        - Month 3 (Mar): 2025 merged C-K, 2026 at L merged L-N
        - ...
        - Month 12 (Dec): 2026 overwrites C, merged C-N (year transition complete)
        """
        year_row = table_cfg["year_row"]
        input_month = int(self.input["report_period"]["date_start"].month)
        input_year = self.input["report_period"]["year"]
    
        # Step 1: Unmerge all merged cells in the year row
        merged_ranges = list(ws.merged_cells.ranges)
        for merged_range in merged_ranges:
            if merged_range.min_row == year_row and merged_range.max_row == year_row:
                ws.unmerge_cells(str(merged_range))
    
        # Step 2: Read current year values and preserve styles
        old_year_cell = ws.cell(row=year_row, column=start_col)
        old_year_in_c = old_year_cell.value
        old_year_style = copy.copy(old_year_cell.font)
        old_year_alignment = copy.copy(old_year_cell.alignment) if old_year_cell.alignment else Alignment(horizontal="center")
    
        # Calculate where the new year would be positioned based on the current month
        # Formula: (end_col - input_month + 2)
        # This represents the column where January of the new year appears
        new_year_position = end_col - input_month + 2
    
        # Read the new year value if it exists at its expected position
        new_year = None
        if new_year_position <= end_col and new_year_position >= start_col:
            potential_new_year = ws.cell(row=year_row, column=new_year_position).value
            # Check if this is actually a different year (year transition)
            if potential_new_year and potential_new_year != old_year_in_c:
                new_year = potential_new_year
    
        # Step 3: Handle year transition (December -> January)
        if input_month == 12:
            # In December, we're about to transition: overwrite old year with new year
            new_year_value = input_year
        
            # Write new year in column C (overwriting old year)
            ws.cell(row=year_row, column=start_col).value = new_year_value
            ws.cell(row=year_row, column=start_col).font = old_year_style
            ws.cell(row=year_row, column=start_col).alignment = old_year_alignment
        
            # Merge across all months (C to N)
            if end_col > start_col:
                ws.merge_cells(
                    start_row=year_row,
                    start_column=start_col,
                    end_row=year_row,
                    end_column=end_col
                )
        
            return  # Year transition complete
    
        # Step 4: Determine year configuration for non-December months
        if input_month == 1:
            # January: First month with new year
            # Old year stays in C, new year appears in N (single cell)
            old_year = old_year_in_c if old_year_in_c else input_year - 1
            new_year_value = input_year
        
            # Old year range: C to M (start_col to end_col - 1)
            old_year_start_col = start_col
            old_year_end_col = end_col - 1
        
            # New year: Only in column N (single cell, no merge)
            new_year_start_col = end_col
            new_year_end_col = end_col
        
        else:
            # February through November
            # The new year needs to shift one column forward each month
            
            if new_year:
                # Step 1: Find where the new year currently is
                # It should be at the position we calculated earlier: new_year_position
                old_new_year_position = new_year_position  # This is where we found it
                
                # Step 2: Calculate where it should move to (one column forward = one column left = -1)
                new_year_start_col = old_new_year_position - 1
                new_year_end_col = end_col
                
                # Step 3: Read the new year value and its style from current position
                new_year_cell = ws.cell(row=year_row, column=old_new_year_position)
                new_year_value = new_year
                new_year_style = copy.copy(new_year_cell.font) if new_year_cell.font else old_year_style
                new_year_alignment = copy.copy(new_year_cell.alignment) if new_year_cell.alignment else old_year_alignment
                
                # Step 4: Clear all cells in the year row
                for col in range(start_col, end_col + 1):
                    ws.cell(row=year_row, column=col).value = None
                
                # Step 5: Write new year at its new position (shifted forward)
                ws.cell(row=year_row, column=new_year_start_col).value = new_year_value
                ws.cell(row=year_row, column=new_year_start_col).font = new_year_style
                ws.cell(row=year_row, column=new_year_start_col).alignment = new_year_alignment
                
                # Old year range: C to before new year
                old_year = old_year_in_c if old_year_in_c else input_year
                old_year_start_col = start_col
                old_year_end_col = new_year_start_col - 1
                
                # Write old year in C
                ws.cell(row=year_row, column=old_year_start_col).value = old_year
                ws.cell(row=year_row, column=old_year_start_col).font = old_year_style
                ws.cell(row=year_row, column=old_year_start_col).alignment = old_year_alignment
            
            else:
                # No year transition yet - only old year across all months
                old_year = old_year_in_c if old_year_in_c else input_year
                old_year_start_col = start_col
                old_year_end_col = end_col
            
                new_year_start_col = None
                new_year_end_col = None
                new_year_value = None
    
        # Step 5: Write year values with styles (only for January and non-transition cases)
        # For February through November with year transition, years are already written above
        if input_month == 1:
            # Write old year in column C
            ws.cell(row=year_row, column=old_year_start_col).value = old_year
            ws.cell(row=year_row, column=old_year_start_col).font = old_year_style
            ws.cell(row=year_row, column=old_year_start_col).alignment = old_year_alignment
        
            # Write new year if it exists
            if new_year_value and new_year_start_col:
                ws.cell(row=year_row, column=new_year_start_col).value = new_year_value
                ws.cell(row=year_row, column=new_year_start_col).font = old_year_style
                ws.cell(row=year_row, column=new_year_start_col).alignment = old_year_alignment
        
        elif not new_year:
            # No year transition yet - write old year only
            ws.cell(row=year_row, column=old_year_start_col).value = old_year
            ws.cell(row=year_row, column=old_year_start_col).font = old_year_style
            ws.cell(row=year_row, column=old_year_start_col).alignment = old_year_alignment
    
        # Step 6: Merge cells
        # Merge old year range
        if old_year_end_col > old_year_start_col:
            ws.merge_cells(
                start_row=year_row,
                start_column=old_year_start_col,
                end_row=year_row,
                end_column=old_year_end_col
            )
    
        # Merge new year range (if applicable and more than one cell)
        if new_year_value and new_year_start_col and new_year_end_col:
            # Format Painter: Copy all styles from old year range to new year range
            # This includes font, fill, alignment, border, number_format, etc.
            old_year_range_size = old_year_end_col - old_year_start_col + 1
            new_year_range_size = new_year_end_col - new_year_start_col + 1
            
            # Copy styles from each cell in old year range to corresponding cell in new year range
            for i in range(new_year_range_size):
                # Source cell from old year range (wrap around if new range is larger)
                source_col = old_year_start_col + (i % old_year_range_size)
                source_cell = ws.cell(row=year_row, column=source_col)
                
                # Target cell in new year range
                target_col = new_year_start_col + i
                target_cell = ws.cell(row=year_row, column=target_col)
                
                # Copy all styles (Format Painter behavior)
                if source_cell.font:
                    target_cell.font = copy.copy(source_cell.font)
                if source_cell.fill:
                    target_cell.fill = copy.copy(source_cell.fill)
                if source_cell.alignment:
                    target_cell.alignment = copy.copy(source_cell.alignment)
                if source_cell.border:
                    target_cell.border = copy.copy(source_cell.border)
                if source_cell.number_format:
                    target_cell.number_format = source_cell.number_format
                if source_cell.protection:
                    target_cell.protection = copy.copy(source_cell.protection)
            
            
            # Apply thick border around the new year range (after copying base styles)
            thick_border = Border(
                left=Side(style='medium'),
                right=Side(style='medium'),
                top=Side(style='thin'),
                bottom=Side(style='medium')
            )
            cell = ws.cell(row=year_row, column=new_year_start_col)
            cell.border = thick_border
            
            # Merge cells if more than one cell
            if new_year_end_col > new_year_start_col:
                ws.merge_cells(
                    start_row=year_row,
                    start_column=new_year_start_col,
                    end_row=year_row,
                    end_column=new_year_end_col
                )
            
         