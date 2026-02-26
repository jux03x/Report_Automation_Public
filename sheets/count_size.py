from sheets.base import BaseSheet
from core.interfaces import SheetProcessor
from openpyxl.styles import Alignment
import sheets.formula

class CountSize(SheetProcessor):

    def process(self):
        # --- read config ---
        cfg = self.config["sheets"]["Count_Size"]

        # --- get worksheet ---
        ws = self.workbook[cfg["sheet_name"]]

        # --- base helper ---
        base = BaseSheet(ws, cfg, self.input)

        # create new column
        new_col = base.copy_last_column()

        # 2. Year 
        #merge
        if int(self.input["report_period"]["date_start"].month) == 1:
            target_cell = ws.cell(row=cfg["year_row"], column=new_col)
            target_cell.value = self.input["report_period"]["year"]
            # Center
            target_cell.alignment = Alignment(horizontal="center")
        elif int(self.input["report_period"]["date_start"].month) == 2:
            ws.merge_cells(
                start_row=cfg["year_row"],
                start_column=new_col - 1,
                end_row=cfg["year_row"],
                end_column=new_col 
            )
        else:
            ws.unmerge_cells(
            start_row=cfg["year_row"],
            start_column=(new_col - int(self.input["report_period"]["date_start"].month)) + 1,
            end_row=cfg["year_row"],
            end_column=new_col - 1
            )
            ws.merge_cells(
                start_row=cfg["year_row"],
                start_column=(new_col - int(self.input["report_period"]["date_start"].month)) + 1,
                end_row=cfg["year_row"],
                end_column=new_col 
            )

        # Month
        ws.cell(
            row=cfg["month_row"],
            column=new_col
        ).value = self.input["report_period"]["date_start"].month

        # fill in data
        ws.cell(
            row=cfg["data"]["document_count"]["row"],
            column=new_col
        ).value = self.input["values"]["document_count"]

        ws.cell(
            row=cfg["data"]["tenant_count"]["row"],
            column=new_col
        ).value = self.input["values"]["tenant_count"]

        ws.cell(
            row=cfg["data"]["storage"]["row"],
            column=new_col
        ).value = self.input["values"]["storage"]

        # Adjust Formular
        src = ws.cell(row=9, column=new_col).value

        if isinstance(src, str) and src.startswith("="):
            ws.cell(row=9, column=new_col).value = sheets.formula.change_formula(src)
        else:
            ws.cell(row=9, column=new_col).value = src


